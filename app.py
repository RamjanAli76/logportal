from flask import Flask, render_template, request, send_file, jsonify
import os
import requests
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as RLImage, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, 'static', 'logo.jpg')

# Unga Google Script Web App URL-ai inga podunga
GAS_URL = "https://script.google.com/macros/s/AKfycbzsZwon1BuwdSLgjJhx5gU6x5wvMHaXMEh9NTvRGyj6Eyy1h1ws1lpSYW9eU3BOpFdS/exec"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    try:
        company = request.form.get('company', 'Service_Report').strip()
        raw_date = request.form.get('date', '')
        
        if raw_date:
            date_obj = datetime.strptime(raw_date, '%Y-%m-%d')
            date_val = date_obj.strftime('%d-%m-%Y')
        else:
            date_val = raw_date

        in_time = f"{request.form.get('in_time')} {request.form.get('in_period')}"
        out_time = f"{request.form.get('out_time')} {request.form.get('out_period')}"
        action_format = request.form.get('format') 
        
        data = {
            'company': company, 'date': date_val, 'works': request.form.get('works'),
            'workers': request.form.get('workers'), 'in_time': in_time,
            'out_time': out_time, 'visit': request.form.get('visit_type'),
            'staff': request.form.get('staff_name'),
            'raw_date': raw_date
        }

        clean_name = company.replace(" ", "_")
        filename_pdf = f"{clean_name}_service_{date_val}.pdf"
        
        # --- AUTO SAVE LOGIC START ---
        # User entha button-ai amukunaalum, background-la PDF generate panni Drive-ku anupum
        temp_pdf_path = generate_pdf(data)
        
        try:
            with open(temp_pdf_path, "rb") as f:
                encoded_string = base64.b64encode(f.read()).decode('utf-8')
                requests.post(GAS_URL, data={
                    'fileData': encoded_string,
                    'filename': filename_pdf,
                    'company': company,
                    'date': raw_date,
                    'mimeType': 'application/pdf'
                }, timeout=5) # 5 seconds-la backend-la upload aagidum
        except Exception as drive_err:
            print(f"Drive Auto-save Error: {drive_err}")
        # --- AUTO SAVE LOGIC END ---

        if action_format == 'excel':
            filename_excel = f"{clean_name}_service_{date_val}.xlsx"
            excel_path = generate_excel(data)
            return send_file(excel_path, as_attachment=True, download_name=filename_excel)
        
        elif action_format == 'drive_excel':
            # Animation-kaga JSON response
            return jsonify({"status": "success", "message": f"Log Sheet Auto-saved in Drive & {company} folder!"})

        else:
            # Direct PDF download
            return send_file(temp_pdf_path, as_attachment=True, download_name=filename_pdf)

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

def generate_excel(data):
    path = os.path.join(BASE_DIR, "temp_report.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:I1')
    ws['A1'] = "SERVICE LOG SHEET"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append(["Date", "Company Name", "Works Carried Out", "Workers", "In Time", "Out Time", "Visit Type", "Staff Name", "Signature"])
    ws.append([data['date'], data['company'], data['works'], data['workers'], data['in_time'], data['out_time'], data['visit'], data['staff'], ""])
  
    widths = {'A': 15, 'B': 20, 'C': 40, 'D': 20, 'E': 10, 'F': 10, 'G': 15, 'H': 20, 'I': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(path)
    return path

def generate_pdf(data):
    path = os.path.join(BASE_DIR, "temp_report.pdf")
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=15, rightMargin=15, topMargin=15, bottomMargin=15)
    styles = getSampleStyleSheet()
    
    style_h = styles["Heading1"]
    style_h.alignment = 1 
    style_h.fontSize = 18
    style_h.spaceAfter = 20
    
    style_n = styles["Normal"]
    style_n.fontSize = 9

    elements = []
    elements.append(Paragraph("<b>SERVICE LOG SHEET</b>", style_h))
    
    logo = RLImage(LOGO_PATH, width=60, height=35) if os.path.exists(LOGO_PATH) else ""
    staff_content = [logo, Paragraph(data['staff'], style_n)]

    table_data = [
        ["Date", "Company Name", "Works Carried Out", "Workers", "Time In/Out", "Visit", "Staff Name", "Signature"],
        [
            data['date'],
            Paragraph(data['company'], style_n), 
            Paragraph(data['works'].replace('\n', '<br/>'), style_n), 
            Paragraph(data['workers'].replace('\n', '<br/>'), style_n),
            f"{data['in_time']}\nto\n{data['out_time']}",
            data['visit'],
            staff_content, 
            ""             
        ]
    ]

    table = Table(table_data, colWidths=[75, 95, 200, 95, 75, 75, 100, 75])
    
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return path

if __name__ == '__main__':
    app.run(debug=True)
